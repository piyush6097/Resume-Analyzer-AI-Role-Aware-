from flask import Flask, request, render_template, jsonify, redirect, url_for, send_file
import os, time, json, unicodedata
import fitz  # PyMuPDF
from analyse_pdf import analyse_resume_st, MODEL_VERSION
from hash_resume import compute_sha256

# --- MySQL data layer ---
from db_mysql import (
    init_db, get_cached_score, save_score,
    add_job_description, list_job_descriptions,
    list_cached_candidates, delete_cached, get_job_description,
    get_roles, get_all_roles_with_skills, get_role_id_by_name,
    upsert_role, upsert_skill, delete_skill
)

# --- XLSX export deps ---
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Absolute paths
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, "uploads")
app.config['EXPORT_FOLDER'] = os.path.join(BASE_DIR, "exports")

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)

# Initialize DB schema
init_db()


@app.route("/health")
def health():
    return "OK", 200


def extract_text_from_resume(pdf_path: str) -> str:
    """
    Basic text extraction via PyMuPDF.
    (OCR fallback can be added later if needed.)
    """
    try:
        doc = fitz.open(pdf_path)
        text = " ".join([page.get_text("text") for page in doc])
        doc.close()
        return text or ""
    except Exception as e:
        print("❌ PDF read error:", e)
        return ""


@app.route("/", methods=["GET", "POST"])
def index():
    results = []
    jds = list_job_descriptions()
    roles = get_roles() or ["Software Developer (Generic)"]  # dropdown populated from DB

    if request.method == "POST":
        resumes = request.files.getlist("resumes")
        job_description = request.form.get("job_description", "").strip()
        selected_jd_id = request.form.get("jd_id")
        role = request.form.get("role")

        if not job_description and not selected_jd_id:
            return render_template("index.html", jds=jds, roles=roles, results=[],
                                   error="Job Description is required.")

        if selected_jd_id and not job_description:
            jd_row = get_job_descriptionsafely(int(selected_jd_id))
        else:
            jd_row = None

        if jd_row:
            jd_text = jd_row[2]
            jd_id_int = int(selected_jd_id)
        else:
            jd_text = job_description
            jd_id_int = None

        for resume_file in resumes:
            filename = resume_file.filename
            if not filename.lower().endswith(".pdf"):
                results.append({
                    "filename": filename,
                    "overall": 0,
                    "cached": False,
                    "summary": "Invalid file: Only PDF allowed",
                    "role": role or ""
                })
                continue

            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            resume_file.save(pdf_path)

            resume_hash = compute_sha256(pdf_path)
            resume_content = extract_text_from_resume(pdf_path)

            analysis = analyse_resume_st(resume_content, jd_text, role or None)
            jd_hash = analysis.get("jd_hash", "")

            cached_score = get_cached_score(
                resume_hash, jd_hash, role or "", MODEL_VERSION
            )

            if cached_score is not None:
                analysis["overall"] = cached_score
                analysis["cached"] = True
            else:
                save_score(
                    resume_hash, jd_hash, role or "",
                    analysis.get("overall", 0),
                    model_version=MODEL_VERSION,
                    jd_id=jd_id_int, source_filename=filename
                )
                analysis["cached"] = False

            results.append({
                "filename": filename,
                "role": role or "",
                "summary": analysis.get("raw_text", ""),
                "details": analysis,
                "overall": analysis.get("overall", 0),
                "cached": analysis.get("cached", False)
            })

    return render_template("index.html", jds=jds, roles=roles, results=results, error=None)

# Helper to guard get_job_description
def get_job_descriptionsafely(jd_id: int):
    try:
        return get_job_description(jd_id)
    except Exception as e:
        print("⚠️ JD fetch error:", e)
        return None


def _join_list(x):
    if x is None:
        return ""
    if isinstance(x, (list, tuple, set)):
        return "; ".join(str(i) for i in x)
    return str(x)


@app.route("/export_xlsx", methods=["POST"])
def export_xlsx():
    # accept full server results payload
    try:
        payload = request.get_json() if request.is_json else json.loads(request.form.get("results", "[]"))
    except Exception as e:
        print("❌ JSON Decode Error (xlsx):", e)
        return "Invalid results data", 400

    if not payload:
        return "No results provided", 400

    timestamp = int(time.time())
    filename = f"results_{timestamp}.xlsx"
    path = os.path.join(app.config['EXPORT_FOLDER'], filename)

    headers = [
        "filename", "role", "overall",
        "technical", "experience", "tools", "soft",
        "matching_core", "matching_tools", "matching_soft",
        "missing_core", "missing_tools", "missing_soft",
        "summary", "cached"
    ]

    def _pick(d, r, key, alt=None):
        if d:
            return d.get(key, d.get(alt) if alt else None)
        return r.get(key, r.get(alt) if alt else None)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # header
        ws.append(headers)
        for c in range(1, len(headers)+1):
            ws.cell(row=1, column=c).font = Font(bold=True)

        # rows
        for r in payload:
            d = r.get("details", {}) if isinstance(r, dict) else {}

            overall = _pick(d, r, "overall")
            technical = _pick(d, r, "technical", "technical_score")
            experience = _pick(d, r, "experience", "experience_score")
            tools = _pick(d, r, "tools", "tools_score")
            soft = _pick(d, r, "soft", "soft_score")

            matching_core = d.get("matching_core", d.get("matching", {}).get("core", r.get("matching_core", []))) if d else r.get("matching_core", "")
            matching_tools = d.get("matching_tools", d.get("matching", {}).get("tools", r.get("matching_tools", []))) if d else r.get("matching_tools", "")
            matching_soft = d.get("matching_soft", d.get("matching", {}).get("soft", r.get("matching_soft", []))) if d else r.get("matching_soft", "")

            missing_core = d.get("missing_core", d.get("missing", {}).get("core", r.get("missing_core", []))) if d else r.get("missing_core", "")
            missing_tools = d.get("missing_tools", d.get("missing", {}).get("tools", r.get("missing_tools", []))) if d else r.get("missing_tools", "")
            missing_soft = d.get("missing_soft", d.get("missing", {}).get("soft", r.get("missing_soft", []))) if d else r.get("missing_soft", "")

            raw_summary = d.get("raw_text") if d else None
            if not raw_summary:
                raw_summary = r.get("summary", "")

            def _fmt(v):
                try:
                    return round(float(v), 2)
                except Exception:
                    return v if v is not None else ""

            row = [
                r.get("filename", ""),
                r.get("role", ""),
                _fmt(overall),
                _fmt(technical),
                _fmt(experience),
                _fmt(tools),
                _fmt(soft),
                _join_list(matching_core),
                _join_list(matching_tools),
                _join_list(matching_soft),
                _join_list(missing_core),
                _join_list(missing_tools),
                _join_list(missing_soft),
                unicodedata.normalize("NFKD", raw_summary or ""),  # keep line breaks
                r.get("cached", False),
            ]
            ws.append(row)

        # wrap text + reasonable widths
        wrap_cols = ["H","I","J","K","L","M","N"]  # match/miss/summary
        for col in wrap_cols:
            for cell in ws[col]:
                cell.alignment = Alignment(wrapText=True, vertical="top")

        # autosize with caps
        maxw = {i: len(h) for i, h in enumerate(headers, start=1)}
        for vals in ws.iter_rows(min_row=2, values_only=True):
            for i, v in enumerate(vals, start=1):
                s = str(v) if v is not None else ""
                maxw[i] = min(max(maxw[i], len(s.split("\n")[0])), 60)

        for i, w in maxw.items():
            ws.column_dimensions[get_column_letter(i)].width = max(10, w + 2)

        wb.save(path)
        print("✅ XLSX export created:", filename)
        return jsonify({"path": f"/download_export/{filename}"})
    except Exception as e:
        print("❌ XLSX export failed:", e)
        return jsonify({"error": "XLSX generation failed", "details": str(e)}), 500


@app.route("/download_export/<filename>")
def download_export(filename):
    path = os.path.join(app.config['EXPORT_FOLDER'], filename)
    if not os.path.exists(path):
        return f"File not found: {filename}", 404
    print("📂 Downloading:", path)
    return send_file(path, as_attachment=True)


# -------- Admin: Job Descriptions --------
@app.route("/admin/jd/add", methods=["POST"])
def admin_add_jd():
    name = request.form.get("name")
    desc = request.form.get("description")
    if not name or not desc:
        return "Name and description required", 400
    add_job_description(name, desc)
    return redirect(url_for("index"))

@app.route("/admin/jd/list", methods=["GET"])
def admin_list_jd():
    return jsonify(list_job_descriptions())


# -------- Admin: Roles/Skills UI --------
@app.route("/admin/skills", methods=["GET"])
def admin_skills_page():
    roles = get_roles()
    role_map = get_all_roles_with_skills()
    return render_template("admin_skills.html", roles=roles, role_map=role_map)

@app.route("/admin/roles/add", methods=["POST"])
def admin_add_role():
    name = (request.form.get("role_name") or "").strip()
    if not name:
        return "Role name required", 400
    upsert_role(name)
    return redirect(url_for("admin_skills_page"))

@app.route("/admin/skills/add", methods=["POST"])
def admin_add_skill():
    role_name = (request.form.get("role_name") or "").strip()
    typ = (request.form.get("type") or "").strip().lower()
    name = (request.form.get("skill_name") or "").strip()
    if not role_name or not typ or not name:
        return "role_name, type, skill_name required", 400
    rid = get_role_id_by_name(role_name)
    if rid is None:
        rid = upsert_role(role_name)
    upsert_skill(rid, name, typ)
    return redirect(url_for("admin_skills_page"))

@app.route("/admin/skills/remove", methods=["POST"])
def admin_remove_skill():
    role_name = (request.form.get("role_name") or "").strip()
    typ = (request.form.get("type") or "").strip().lower()
    name = (request.form.get("skill_name") or "").strip()
    rid = get_role_id_by_name(role_name)
    if rid is None:
        return "Role not found", 404
    delete_skill(rid, name, typ)
    return redirect(url_for("admin_skills_page"))

@app.route("/admin/candidates", methods=["GET"])
def admin_list_candidates():
    return jsonify(list_cached_candidates())

@app.route("/admin/candidate/delete", methods=["POST"])
def admin_delete_candidate():
    resume_hash = request.form.get("resume_hash")
    jd_hash = request.form.get("jd_hash")
    role = request.form.get("role")
    delete_cached(resume_hash, jd_hash, role)
    return "deleted", 200


if __name__ == "__main__":
    print("📁 EXPORT_FOLDER:", app.config['EXPORT_FOLDER'])
    print("📁 UPLOAD_FOLDER:", app.config['UPLOAD_FOLDER'])
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
